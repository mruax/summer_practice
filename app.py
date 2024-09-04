import os
import sys
from math import ceil, floor, isclose
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns
from PySide6 import QtGui, QtWidgets
from PySide6.QtGui import QPixmap, Qt, QAction, QFont
from PySide6.QtWidgets import QWidget, QVBoxLayout, QLabel, QFileDialog, QMessageBox
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from ui import Ui_MainWindow


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.x = []
        self.v = []
        self.t = []
        self.e = []

        self.N = 0
        self.M = 0

        self.energy_flag = False
        self.points = 0

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.icon = QtGui.QIcon(QtGui.QPixmap(Path("src/logo.ico")))
        self.setWindowIcon(self.icon)

        self.input_fields = [self.ui.lineEdit_t0, self.ui.lineEdit_t1,
                             self.ui.lineEdit_v0, self.ui.lineEdit_x0,
                             self.ui.lineEdit_gamma, self.ui.lineEdit_omega0,
                             self.ui.lineEdit_m, self.ui.lineEdit_a0, self.ui.lineEdit_OMEGA,
                             self.ui.lineEdit_dt1, self.ui.lineEdit_dt2]

        for field in self.input_fields:
            field.textEdited.connect(lambda _, f=field: self.reset_style(f))

        self.ui.pushButton_graph.clicked.connect(self.print_graph)
        self.ui.pushButton_reset.clicked.connect(self.reset)
        self.ui.pushButton_print.clicked.connect(self.excel_graph)
        self.ui.pushButton_fileData.clicked.connect(self.file_data)

        self.qpushbutton_disabled_style = """
            QPushButton:disabled {
                background-color: #d3d3d3;
                color: #808080;
                border: 1px solid #a9a9a9;
            }
        """

        self.current_graph = False
        self.ui.pushButton_left_arrow.setDisabled(True)
        self.ui.pushButton_left_arrow.setStyleSheet(self.qpushbutton_disabled_style)
        self.ui.pushButton_right_arrow.setDisabled(True)
        self.ui.pushButton_right_arrow.setStyleSheet(self.qpushbutton_disabled_style)

        self.ui.pushButton_left_arrow.clicked.connect(self.prev_graph)
        self.ui.pushButton_right_arrow.clicked.connect(self.next_graph)

        self.ui.pushButton_print.setDisabled(True)
        self.ui.pushButton_print.setStyleSheet(self.qpushbutton_disabled_style)

        menu_bar = self.menuBar()
        menu = menu_bar.addMenu("Дополнительно")

        open_action1 = QAction("Ввод данных", self)
        open_action1.triggered.connect(self.input_data_window)
        menu.addAction(open_action1)

        open_action2 = QAction("Помощь", self)
        open_action2.triggered.connect(self.help_window)
        menu.addAction(open_action2)

        open_action3 = QAction("О программе", self)
        open_action3.triggered.connect(self.info_window)
        menu.addAction(open_action3)

    def resizeEvent(self, event):
        try:
            self.pixmap = QPixmap(Path(f"src/graph{self.current_graph}.png"))
            graph_size = self.ui.label_graph.size()
            scaled_pixmap = self.pixmap.scaled(graph_size.width(), graph_size.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.ui.label_graph.setPixmap(scaled_pixmap)
            self.ui.label_graph.setMinimumSize(1, 1)
        except Exception:
            pass
        super().resizeEvent(event)

    def print_graph(self, event=None):
        if self.check():
            self.reset_colour()

            try:
                self.create_plot()
                pixmap = self.plot_to_label()
                self.ui.label_graph.setPixmap(pixmap)
            except Exception as e:
                self.show_error_message(e)

    def input_data_window(self, event=None):
        self.new_window4 = QWidget()
        self.new_window4.setGeometry(360, 220, 200, 200)
        self.new_window4.setFixedSize(480, 400)
        self.new_window4.setWindowTitle("Ввод данных")

        self.new_window4.setWindowIcon(self.icon)

        layout = QVBoxLayout()

        label_2 = QLabel("Ввод данных")
        font = QFont()
        font.setFamilies([u"Comic Sans MS"])
        font.setPointSize(16)
        label_2.setFont(font)
        label_2.setAlignment(Qt.AlignCenter)

        label = QLabel("Для ввода данных из файла можно использовать\n"
                       "один из форматов:\n"
                       "txt, csv, xlsx\n"
                       "Пример входных данных для .txt (через ;):\n"
                       "t0;t1;v0;x0;gamma;w0;m;A0;OMEGA;dt1;dt2\n"
                       "0;20;5;-5;1;1;1;1;1;0.5;0.25\n"
                       "Пример для .csv (через ;):\n"
                       "t0;t1;v0;x0;gamma;w0;m;A0;OMEGA;dt1;dt2\n"
                       "0;10;1;1;1;1;1;1;1;0.5;0.25\n"
                       "Пример для .xlsx (через табуляцию):\n"
                       "t0 t1 v0 x0 gamma w0 m A0 OMEGA dt1 dt2\n"
                       "0 25 -3 10 1 1 4 1,2 1 0,5 0,25")
        font1 = QFont()
        font1.setFamilies([u"Comic Sans MS"])
        font1.setPointSize(14)
        label.setFont(font1)
        label.setAlignment(Qt.AlignLeft)

        layout.addWidget(label_2)
        layout.addWidget(label)
        self.new_window4.setLayout(layout)

        self.new_window4.show()

    def info_window(self, event=None):
        self.new_window2 = QWidget()
        self.new_window2.setGeometry(360, 220, 200, 200)
        self.new_window2.setFixedSize(360, 220)
        self.new_window2.setWindowTitle("О программе")

        self.new_window2.setWindowIcon(self.icon)

        layout = QVBoxLayout()

        label_2 = QLabel("О программе")
        font = QFont()
        font.setFamilies([u"Comic Sans MS"])
        font.setPointSize(16)
        label_2.setFont(font)
        label_2.setAlignment(Qt.AlignCenter)

        label = QLabel("Работу выполнил: Черняков Матвей\n"
                       "Студент группы Фт-220008\n"
                       "Руководитель практики:\n"
                       "Кибардин Алексей Владимирович")
        font1 = QFont()
        font1.setFamilies([u"Comic Sans MS"])
        font1.setPointSize(14)
        label.setFont(font1)
        label.setAlignment(Qt.AlignLeft)

        layout.addWidget(label_2)
        layout.addWidget(label)
        self.new_window2.setLayout(layout)

        self.new_window2.show()

    def help_window(self, event=None):
        self.new_window3 = QWidget()
        self.new_window3.setGeometry(360, 220, 200, 200)
        self.new_window3.setFixedSize(625, 560)
        self.new_window3.setWindowTitle("Помощь")

        self.new_window3.setWindowIcon(self.icon)

        layout = QVBoxLayout()

        label_2 = QLabel("Помощь")
        font = QFont()
        font.setFamilies([u"Comic Sans MS"])
        font.setPointSize(16)
        label_2.setFont(font)
        label_2.setAlignment(Qt.AlignCenter)

        label = QLabel("Входные данные (все представлены в единицах СИ):\n"
                       "t0,t1 – начальный и конечный момент времени расчетов (с);\n"
                       "v0 – начальная скорость (м/с);\n"
                       "x0 – начальное положение груза (м);\n"
                       "γ – коэффициент трения;\n"
                       "ω0 – собственная частота колебаний осциллятора (рад/с);\n"
                       "dt1 – период расчетов или же интервал времени,\n"
                       "через который выдается текущее значение переменных (с);\n"
                       "dt2 (∆t) – параметр дискретизации (с);\n"
                       "Гармоническая вынуждающая сила, представленная выражением\n"
                       "F(t) = A0 ∙ cos(Ωt),\n"
                       "где Ω – собственная частота вынуждающей силы (рад/с), A0 – константа (Н);\n"
                       "m – масса груза (кг).\n"
                       "Корректность вводимых данных:\n"
                       "t0 < t1; t0 + dt1 < t1\n"
                       "dt1 > 0; dt2 > 0; dt1 >= dt2\n"
                       "γ >= 0; ω0 >= 0; m >= 0; Ω >= 0\n"
                       "Полная энергия системы считается при нулевых внешних силах:\n"
                       "A0 = 0; Ω = 0; γ = 0\n"
                       "Результаты измерений вносятся в таблицу excel.")
        font1 = QFont()
        font1.setFamilies([u"Comic Sans MS"])
        font1.setPointSize(12)
        label.setFont(font1)
        label.setAlignment(Qt.AlignLeft)

        layout.addWidget(label_2)
        layout.addWidget(label)
        self.new_window3.setLayout(layout)

        self.new_window3.show()

    def create_plot(self):
        # Параметры
        t0 = float(self.input_fields[0].text())      # Начальное время
        t1 = float(self.input_fields[1].text())      # Конечное время
        v0 = float(self.input_fields[2].text())      # Начальная скорость
        x0 = float(self.input_fields[3].text())      # Начальное положение груза
        gamma = float(self.input_fields[4].text())   # Коэффициент трения
        omega0 = float(self.input_fields[5].text())  # Собственная частота колебаний осциллятора
        m = float(self.input_fields[6].text())       # Масса груза
        A0 = float(self.input_fields[7].text())      # Константа гармонической вынуждающей силы (Амплитуда силы)
        OMEGA = float(self.input_fields[8].text())   # Собственная частота вынуждающей силы
        dt1 = float(self.input_fields[9].text())     # Период расчетов
        dt2 = float(self.input_fields[10].text())    # Параметр дискретизации

        N = ceil((t1 - t0) / dt1)
        M = ceil(dt1 / dt2)

        self.N = N
        self.M = M

        index = 1
        if A0 == 0 or OMEGA == 0:
            OMEGA_period = 0
        else:
            OMEGA_period = 1 / OMEGA

        energy_flag = True if gamma == 0 and (A0 == 0 or OMEGA == 0) else False
        e = list(False for _ in range(0, N * M + 2, 1))
        self.energy_flag = energy_flag

        # Инициализация массивов для скорости и положения
        t = list(False for _ in range(0, N * M + 2, 1))
        v = list(False for _ in range(0, N * M + 2, 1))
        x = list(False for _ in range(0, N * M + 2, 1))

        # Начальные условия
        t[0] = t0
        v[0] = v0
        x[0] = x0

        if energy_flag:
            e[0] = (m * v[0] ** 2 + m * omega0 ** 2 * x[0] ** 2) / 2
            self.points = N

        # Итеративное вычисление t, v, x
        for i in range(1, N + 1):
            for k in range(1, M + 1):
                if t[index - 1] + dt2 > t0 + dt1 * i:
                    t[index] = t0 + dt1 * i
                else:
                    t[index] = t[index - 1] + dt2
                if t[index] > t1:
                    t[index] = t1

                if A0 == 0 or OMEGA == 0:
                    driving_force = 0
                else:
                    driving_force = A0 * np.cos(OMEGA * t[index]) / m if abs(t[index] / OMEGA_period - floor(t[index] / OMEGA_period)) < dt2 else 0
                v[index] = v[index - 1] + (-omega0 ** 2 * x[index - 1] - gamma * v[index - 1] + driving_force) * dt2
                x[index] = x[index - 1] + v[index] * dt2

                if energy_flag:
                    if isclose(t[index], t0 + dt1 * i, rel_tol=1e-05):
                        e[index] = (m * v[index] ** 2 + m * omega0 ** 2 * x[index] ** 2) / 2
                if t[index] == t1:
                    break
                index += 1

        # Вычисление пограничного финального значения времени (при необходимости)
        if t1 not in t:
            while t[-1] is False:
                del t[-1]
                del v[-1]
                del x[-1]
                if energy_flag:
                    del e[-1]
            t.append(t1)
            if A0 == 0 or OMEGA == 0:
                driving_force = 0
            else:
                driving_force = A0 * np.cos(OMEGA * t[-1]) / m if abs(t[-1] / OMEGA_period - floor(t[-1] / OMEGA_period)) < dt2 else 0
            v.append(v[-1] + (-omega0 ** 2 * x[-1] - gamma * v[-1] + driving_force) * dt2)
            x.append(x[-1] + v[-1] * dt2)
            if energy_flag:
                e.append((m * v[-1] ** 2 + m * omega0 ** 2 * x[-1] ** 2) / 2)
        else:
            count = t.count(t1)
            if count == 1:
                while t[-1] != t1:
                    del t[-1]
                    del v[-1]
                    del x[-1]
                    if energy_flag:
                        del e[-1]
            else:
                flag = True
                while flag:
                    if t[-1] == t1:
                        count -= 1

                    del t[-1]
                    del v[-1]
                    del x[-1]
                    if energy_flag:
                        del e[-1]

                    if count == 1:
                        flag = False

        try:
            self.x = np.asarray(x, dtype=np.float32)
            self.v = np.asarray(v, dtype=np.float32)
            self.t = np.asarray(t, dtype=np.float32)
            if any(np.isinf(self.x)) or any(np.isinf(self.v)):
                raise Exception('Входные параметры приводят к слишком большим числам!')
        except Exception as e:
            self.show_error_message(e)
            return
        if energy_flag:
            self.e = np.asarray([data if data else -1 for data in e], dtype=np.float32)

        # Построение графиков результатов
        sns.set(style="whitegrid")
        plt.figure(figsize=(14, 7))
        plt.plot(t, x, label='Положение $x(t)$')
        plt.plot(t, v, label='Скорость $v(t)$', color='orange')
        plt.title('Затухающий гармонический осциллятор с внешней силой')
        plt.xlabel('Время $t$')
        plt.ylabel('Положение $x(t)$ и Скорость $v(t)$')
        plt.legend()
        plt.savefig(Path("src/graph1.png"), format='png', bbox_inches='tight')
        plt.close()

        sns.set(style="whitegrid")
        plt.figure(figsize=(14, 7))
        plt.plot(t, x, label='Положение $x(t)$')
        plt.title('Затухающий гармонический осциллятор с внешней силой')
        plt.xlabel('Время $t$')
        plt.ylabel('Положение $x(t)$')
        plt.legend()
        plt.savefig(Path("src/graph2.png"), format='png', bbox_inches='tight')
        plt.close()

        sns.set(style="whitegrid")
        plt.figure(figsize=(14, 7))
        plt.plot(t, v, label='Скорость $v(t)$', color='orange')
        plt.title('Затухающий гармонический осциллятор с внешней силой')
        plt.xlabel('Время $t$')
        plt.ylabel('Скорость $v(t)$')
        plt.legend()
        plt.savefig(Path("src/graph3.png"), format='png', bbox_inches='tight')
        plt.close()

        self.ui.pushButton_print.setStyleSheet("")
        self.ui.pushButton_print.setEnabled(True)

        self.current_graph = 1
        self.ui.pushButton_right_arrow.setStyleSheet("")
        self.ui.pushButton_right_arrow.setEnabled(True)

    def excel_graph(self):
        excel_filename = Path('src/data.xlsx')
        if self.energy_flag:
            data = {'t': self.t, 'x': self.x, 'v': self.v, "Полная энергия замкнутой системы": self.e}
        else:
            data = {'t': self.t, 'x': self.x, 'v': self.v}
        try:
            df = pd.DataFrame(data)
            df.to_excel(excel_filename, index=False)
        except Exception as e:
            self.show_error_message(e)
            return

        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        last_row = ws.max_row
        if self.energy_flag:  # Расчет при замкнутости системы
            for cell in range(2, last_row + 1):
                if (cell - 2) % self.M != 0:
                    ws[f"D{cell}"].value = ""

            ws[f'E1'] = "Средняя составляющая энергии"
            ws[f'E2'] = f"=AVERAGE(D2:D{last_row})"

            ws[f'F1'] = "Относительная погрешность"
            for row in range(2, last_row + 1):
                if (row - 2) % self.M == 0:
                    ws[f'F{row}'] = f"=ABS($E$2 - $D{row}) / ABS($E$2)"

            ws[f'G1'] = "Количество рассчетных точек"
            ws[f'G2'] = self.points

            ws[f'H1'] = "Среднеквадратичная относительная погрешность"
            ws[f'H2'] = f"=SQRT(1 / $G$2 * SUMPRODUCT(($F$2:$F${last_row})^2))"

        # Оформление таблицы
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 4
            ws.column_dimensions[column].width = adjusted_width
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        fill1 = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        fill2 = PatternFill(start_color='E4E4E4', end_color='E4E4E4', fill_type='solid')
        fill3 = PatternFill(start_color='C5C6D5', end_color='C5C6D5', fill_type='solid')
        first_row_flag = True
        for row in ws.iter_rows():
            for cell in row:
                if first_row_flag:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    cell.fill = fill1
                else:
                    cell.fill = fill2
                cell.border = thin_border
            if first_row_flag:
                first_row_flag = False
        for row in range(2, last_row + 1):
            if (row - 2) % self.M == 0:
                ws[f'A{row}'].fill = fill3
        if self.energy_flag:
            ws[f'G2'].fill = fill3
        wb.save(excel_filename)
        wb.close()
        os.startfile(excel_filename)

    def file_data(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Open File",
            "",
            "CSV Files (*.csv);;Text Files (*.txt);;Excel Files (*.xlsx)"
        )

        if file_name:
            file_path = Path(file_name)
            file_extension = os.path.splitext(file_name)[1]
            if file_extension == ".csv" or file_extension == ".txt":
                with open(file_path, "r") as file:
                    file.readline()
                    line_input = file.readline().split(";")
                    for id, line in enumerate(self.input_fields):
                        try:
                            line.setText(line_input[id])
                        except Exception:
                            pass
            elif file_extension == ".xlsx":
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                for id, line in enumerate(self.input_fields):
                    try:
                        line.setText(str(data[1][id]))
                    except Exception:
                        pass
                workbook.close()

    def next_graph(self):
        self.current_graph += 1
        self.ui.pushButton_left_arrow.setStyleSheet("")
        self.ui.pushButton_left_arrow.setEnabled(True)

        if self.current_graph == 2:
            self.ui.pushButton_right_arrow.setStyleSheet("")
            self.ui.pushButton_right_arrow.setEnabled(True)
        else:
            self.ui.pushButton_right_arrow.setDisabled(True)
            self.ui.pushButton_right_arrow.setStyleSheet(self.qpushbutton_disabled_style)

        self.ui.label_graph.setPixmap(QPixmap())
        pixmap = self.plot_to_label()
        self.ui.label_graph.setPixmap(pixmap)

    def prev_graph(self):
        self.current_graph -= 1
        self.ui.pushButton_right_arrow.setStyleSheet("")
        self.ui.pushButton_right_arrow.setEnabled(True)

        if self.current_graph == 2:
            self.ui.pushButton_left_arrow.setStyleSheet("")
            self.ui.pushButton_left_arrow.setEnabled(True)
        else:
            self.ui.pushButton_left_arrow.setDisabled(True)
            self.ui.pushButton_left_arrow.setStyleSheet(self.qpushbutton_disabled_style)

        self.ui.label_graph.setPixmap(QPixmap())
        pixmap = self.plot_to_label()
        self.ui.label_graph.setPixmap(pixmap)

    def plot_to_label(self):
        pixmap = QPixmap(Path(f"src/graph{self.current_graph}.png"))
        graph_size = self.ui.label_graph.size()
        scaled_pixmap = pixmap.scaled(graph_size.width(), graph_size.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        return scaled_pixmap

    def check(self):
        flag = False
        styleSheet = "color: red"
        temp = []
        for line in self.input_fields:
            line_text = line.text()
            line_text = line_text.replace(',', '.')
            temp.append(line_text)
            if line_text == "":
                line.setText("null")
                line.setStyleSheet(styleSheet)
                flag = True
            else:
                try:  # Только числа
                    _ = float(line_text)
                except ValueError:
                    line.setStyleSheet(styleSheet)
                    flag = True

        for i in range(len(temp)):
            self.input_fields[i].setText(temp[i])

        # Проверка корректности значений
        try:  # Верный промежуток времени
            if float(temp[0]) >= float(temp[1]) or float(temp[0]) + float(temp[10]) >= float(temp[1]):
                flag = True
                self.input_fields[0].setStyleSheet(styleSheet)
                self.input_fields[1].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[0].setStyleSheet(styleSheet)
            self.input_fields[1].setStyleSheet(styleSheet)
        try:  # Корректный период расчета и параметр дискретизации
            if (float(temp[9]) <= 0 or float(temp[10]) <= 0) or (float(temp[10]) > float(temp[9])):
                flag = True
                self.input_fields[9].setStyleSheet(styleSheet)
                self.input_fields[10].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[9].setStyleSheet(styleSheet)
            self.input_fields[10].setStyleSheet(styleSheet)
        try:  # Корректный коэффициент трения
            if float(temp[4]) < 0:
                flag = True
                self.input_fields[4].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[4].setStyleSheet(styleSheet)
        try:  # Корректная собственная частота колебаний осциллятора
            if float(temp[5]) < 0:
                flag = True
                self.input_fields[5].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[5].setStyleSheet(styleSheet)
        try:  # Корректная масса груза
            if float(temp[6]) < 0:
                flag = True
                self.input_fields[6].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[6].setStyleSheet(styleSheet)
        try:  # Корректная собственная частота вынуждающей силы
            if float(temp[8]) < 0:
                flag = True
                self.input_fields[8].setStyleSheet(styleSheet)
        except ValueError:
            flag = True
            self.input_fields[8].setStyleSheet(styleSheet)

        return not flag

    def reset(self, event=None):
        self.reset_text()
        self.reset_colour()
        self.ui.label_graph.setPixmap(QPixmap())

        self.ui.pushButton_print.setDisabled(True)
        self.ui.pushButton_print.setStyleSheet(self.qpushbutton_disabled_style)

        self.current_graph = False
        self.ui.pushButton_left_arrow.setDisabled(True)
        self.ui.pushButton_left_arrow.setStyleSheet(self.qpushbutton_disabled_style)
        self.ui.pushButton_right_arrow.setDisabled(True)
        self.ui.pushButton_right_arrow.setStyleSheet(self.qpushbutton_disabled_style)

    def reset_text(self):
        for line in self.input_fields:
            line.setText("")

    def reset_colour(self):
        styleSheet = "color: black"
        for line in self.input_fields:
            line.setStyleSheet(styleSheet)

    def reset_style(self, line):
        styleSheet = "color: black"
        line.setStyleSheet(styleSheet)

    def show_error_message(self, error_text):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setWindowTitle("Ошибка")
        msg_box.setText(str(error_text))
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec()


if __name__ == "__main__":
    app = QtWidgets.QApplication([])

    window = MainWindow()
    window.resize(1000, 600)
    window.show()

    window.setWindowTitle("Гармонический осциллятор v1.0")

    icon = QtGui.QIcon(QtGui.QPixmap(Path("src/logo.ico")))
    window.setWindowIcon(icon)

    sys.exit(app.exec())
